#!/usr/bin/env python3
"""
LaptopSync Server - Run this on your Windows PC.
Receives laptop test data from test laptops and logs to Daily_Log.xlsx.

Requirements:
    pip install flask openpyxl

Folder structure (auto-created in your Documents folder):
    Documents/
    └── LaptopSync/
        ├── Laptop_Templates.xlsx   ← put your template here
        └── Daily_Log.xlsx          ← auto-created
"""

import os
import re
import sys
from datetime import datetime
from collections import Counter
from pathlib import Path

from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

# ---------------------------------------------------
# PATHS
# ---------------------------------------------------

DOCUMENTS_DIR = Path.home() / "Documents" / "LaptopSync"
TEMPLATE_FILE = DOCUMENTS_DIR / "Laptop_Templates.xlsx"
DAILY_FILE    = DOCUMENTS_DIR / "Daily_Log.xlsx"
PORT          = 5050

# ---------------------------------------------------
# INIT
# ---------------------------------------------------

app = Flask(__name__)

def setup_directory():
    DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
    if not TEMPLATE_FILE.exists():
        print(f"\n  ERROR: Template file not found.")
        print(f"  Please place 'Laptop_Templates.xlsx' in:")
        print(f"  {DOCUMENTS_DIR}\n")
        sys.exit(1)
    print(f"  Template : {TEMPLATE_FILE}")
    print(f"  Log file : {DAILY_FILE}")

def initialize_daily_log():
    """Create Daily_Log.xlsx from template headers if it doesn't exist."""
    if DAILY_FILE.exists():
        return

    template_wb = load_workbook(TEMPLATE_FILE)
    template_ws = template_wb.active

    new_wb = Workbook()
    new_ws = new_wb.active

    # Copy header row only
    for col_idx, cell in enumerate(template_ws[1], start=1):
        new_ws.cell(row=1, column=col_idx, value=cell.value)

    new_wb.save(DAILY_FILE)
    print(f"  Created new daily log: {DAILY_FILE}")

# ---------------------------------------------------
# CPU PARSER
# ---------------------------------------------------

def parse_cpu_info(raw):
    """
    Parse a /proc/cpuinfo model name string into structured fields.

    Returns:
        cpu_type   (str) -> column AP  e.g. "Intel Core i5", "Intel Core Ultra 7", "AMD Ryzen 5"
        cpu_series (str) -> column S   e.g. "8250U", "155H", "N4020", "5600U"
        cpu_freq   (str) -> column AO  e.g. "1.60GHz" or "Unknown"

    Covers Intel and AMD laptop CPUs from roughly 2010 to present:
        Intel  : Core Ultra (5/7/9), Core i-series (i3-i9), Core m-series (m3/m5/m7),
                 Celeron (N/J/numeric), Pentium (Silver/Gold/legacy), Atom
        AMD    : Ryzen AI, Ryzen (3/5/7/9 + PRO), A-series APU, E-series,
                 Athlon (Silver/Gold), FX
    """
    s          = raw.strip()
    cpu_type   = "Unknown"
    cpu_series = "Unknown"

    # ----------------------------------------------------------
    # INTEL
    # ----------------------------------------------------------

    # Intel Core Ultra (Meteor Lake / Arrow Lake / Lunar Lake 2023+)
    # e.g. "Intel(R) Core(TM) Ultra 5 125U"
    #      "Intel(R) Core(TM) Ultra 7 155H"
    #      "Intel(R) Core(TM) Ultra 7 Pro 265H"
    m = re.search(r"Core\s*(?:\(TM\)\s*)?Ultra\s+([579])\s+(?:Pro\s+)?(\d{3}\w*)", s, re.IGNORECASE)
    if m:
        cpu_type   = f"Intel Core Ultra {m.group(1)}"
        cpu_series = m.group(2)

    # Intel Core i-series (Sandy Bridge 2011 -> present)
    # e.g. "Core(TM) i5-8250U"  "Core i7-1165G7"  "Core i9-13900H"  "Core i5-2520M"
    elif m := re.search(r"Core\s*(?:\(TM\)\s*)?(i[3579])-(\d{4,5}\w*)", s, re.IGNORECASE):
        cpu_type   = f"Intel Core {m.group(1).lower()}"
        cpu_series = m.group(2)

    # Intel Core m-series (Broadwell-Y / Kaby Lake-Y 2015-2019)
    # e.g. "Core(TM) m3-8100Y"  "Core m5-6Y54"  "Core m7-6Y75"
    elif m := re.search(r"Core\s*(?:\(TM\)\s*)?m([357])-(\d+\w+)", s, re.IGNORECASE):
        cpu_type   = f"Intel Core m{m.group(1)}"
        cpu_series = m.group(2)

    # Intel Celeron -- N/J-series (Bay Trail through Jasper Lake 2013-2021)
    # e.g. "Celeron(R) N4020"  "Celeron N2840"  "Celeron N5100"  "Celeron J4125"
    # Tolerates optional "(R)" and "CPU" token before the model number
    elif m := re.search(r"Celeron\s*(?:\(R\)\s*)?(?:CPU\s+)?([NJ]\d{4})", s, re.IGNORECASE):
        cpu_type   = "Intel Celeron"
        cpu_series = m.group(1).upper()

    # Intel Celeron -- older numeric / alphanumeric models (Sandy/Ivy Bridge era 2011-2013)
    # e.g. "Celeron CPU 847"  "Celeron CPU 2955U"  "Celeron 1000M"
    elif m := re.search(r"Celeron\s*(?:\(R\)\s*)?(?:CPU\s+)?(\d{3,4}[A-Z]*)", s, re.IGNORECASE):
        cpu_type   = "Intel Celeron"
        cpu_series = m.group(1).upper()

    # Intel Pentium Silver / Gold (Gemini Lake / Kaby Lake refresh / Tiger Lake 2017-2022)
    # e.g. "Pentium(R) Silver N5000"  "Pentium(R) Gold 4415U"  "Pentium(R) Gold 7505"
    elif m := re.search(
        r"Pentium\s*(?:\(R\)\s*)?(?:(Silver|Gold)\s+)?(?:CPU\s+)?([NJ]?\d{3,5}\w*)",
        s, re.IGNORECASE
    ):
        grade      = " " + m.group(1).title() if m.group(1) else ""
        cpu_type   = f"Intel Pentium{grade}"
        cpu_series = m.group(2).upper()

    # Intel Pentium -- older letter-prefixed models (Sandy/Ivy Bridge 2011-2013)
    # e.g. "Pentium CPU B960"  "Pentium CPU 2020M"
    elif m := re.search(
        r"Pentium\s*(?:\(R\)\s*)?(?:CPU\s+)?([A-Z]\d{3,4}[A-Z]*|\d{4}[A-Z])",
        s, re.IGNORECASE
    ):
        cpu_type   = "Intel Pentium"
        cpu_series = m.group(1).upper()

    # Intel Atom (Bay Trail / Cherry Trail / Apollo Lake 2013-2017)
    # e.g. "Atom(TM) x5-Z8350"  "Atom(TM) CPU Z3735F"  "Atom x7-Z8750"
    elif m := re.search(
        r"Atom\s*(?:\(TM\)\s*)?(?:x[357]-|CPU\s+)?([A-Z]\d{4,5}\w*)",
        s, re.IGNORECASE
    ):
        cpu_type   = "Intel Atom"
        cpu_series = m.group(1).upper()

    # ----------------------------------------------------------
    # AMD
    # ----------------------------------------------------------

    # AMD Ryzen AI (Strix Point / Hawk Point 2024+)
    # e.g. "Ryzen AI 5 340"  "Ryzen AI 7 Pro 360"  "Ryzen AI 9 HX 370"
    elif m := re.search(r"Ryzen\s+AI\s+([3579])\s+(?:(?:Pro|HX)\s+)?(\d{3}\w*)", s, re.IGNORECASE):
        cpu_type   = f"AMD Ryzen AI {m.group(1)}"
        cpu_series = m.group(2)

    # AMD Ryzen (Zen 1 2018 -> present), including PRO variants
    # e.g. "Ryzen 5 3500U"  "Ryzen 7 5700U"  "Ryzen 9 5900HX"  "Ryzen 5 PRO 4650U"
    elif m := re.search(r"Ryzen\s+([3579])\s+(?:Pro\s+)?(\d{4}\w*)", s, re.IGNORECASE):
        cpu_type   = f"AMD Ryzen {m.group(1)}"
        cpu_series = m.group(2)

    # AMD A-series APU (Carrizo / Bristol Ridge / Stoney Ridge 2014-2018)
    # e.g. "A6-9220"  "A10-9620P"  "A4-9125"  "A12-9720P"
    elif m := re.search(r"\bA(\d{1,2})-(\d{4}\w*)", s, re.IGNORECASE):
        cpu_type   = f"AMD A{m.group(1)}"
        cpu_series = m.group(2)

    # AMD E-series (Brazos / Beema / Stoney ultra-budget 2011-2018)
    # e.g. "E1-6010"  "E2-9000e"
    elif m := re.search(r"\bE(\d)-(\d{3,4}\w*)", s, re.IGNORECASE):
        cpu_type   = f"AMD E{m.group(1)}"
        cpu_series = m.group(2)

    # AMD Athlon Silver / Gold (Raven Ridge / Picasso budget 2019-2021)
    # e.g. "Athlon Silver 3050U"  "Athlon Gold 3150U"
    elif m := re.search(r"Athlon\s+(?:(Silver|Gold)\s+)?(\d{4}\w*)", s, re.IGNORECASE):
        grade      = " " + m.group(1).title() if m.group(1) else ""
        cpu_type   = f"AMD Athlon{grade}"
        cpu_series = m.group(2)

    # AMD FX (Kaveri / Carrizo FX branding 2014-2017)
    # e.g. "FX-7500"  "FX-9800P"
    elif m := re.search(r"\bFX-(\d{4}\w*)", s, re.IGNORECASE):
        cpu_type   = "AMD FX"
        cpu_series = m.group(1)

    # ----------------------------------------------------------
    # Frequency (universal)
    # ----------------------------------------------------------
    freq_m   = re.search(r"@\s*([\d.]+\s*GHz)", s, re.IGNORECASE)
    cpu_freq = freq_m.group(1).replace(" ", "") if freq_m else "Unknown"

    return cpu_type, cpu_series, cpu_freq

# ---------------------------------------------------
# RAM / MISC HELPERS
# ---------------------------------------------------

def parse_ram_info(ram_slots):
    sizes    = []
    ram_type = "Unknown"

    for slot in ram_slots:
        if not slot:
            continue
        size_match = re.search(r"(\d+)\s*GB", slot, re.IGNORECASE)
        if size_match:
            sizes.append(int(size_match.group(1)))
        type_match = re.search(r"(DDR\d)", slot, re.IGNORECASE)
        if type_match:
            ram_type = type_match.group(1).upper()

    if not sizes:
        return "Unknown", "0GB", "Unknown"

    size_counts = Counter(sizes)
    if len(set(sizes)) == 1:
        ram_config = f"{len(sizes)}x{sizes[0]}GB"
    else:
        ram_config = " + ".join(f"{qty}x{size}GB" for size, qty in size_counts.items())

    return ram_config, f"{sum(sizes)}GB", ram_type


def clean_total_ram(ram_config):
    total = sum(int(qty) * int(size) for qty, size in re.findall(r"(\d+)x(\d+)GB", ram_config))
    return f"{total}GB"


def append_to_log(data):
    """Find matching template row and append to daily log."""
    template_wb = load_workbook(TEMPLATE_FILE)
    template_ws = template_wb.active

    daily_wb = load_workbook(DAILY_FILE)
    daily_ws = daily_wb.active

    model = data.get("model", "")

    # Find matching template row.
    # Checks both directions so that a sent model of "HP EliteBook 840 G1"
    # matches a template entry of "EliteBook 840 G1", and vice-versa.
    template_row = None
    sent_lower   = model.strip().lower()
    for row in template_ws.iter_rows(min_row=2):
        template_model = str(row[5].value).strip().lower()
        if sent_lower in template_model or template_model in sent_lower:
            template_row = row
            break

    if not template_row:
        raise ValueError(f"Model '{model}' not found in template sheet")

    # Build new row from template
    new_row = [cell.value for cell in template_row]
    while len(new_row) < 42:
        new_row.append(None)

    # Parse CPU -- single call returns all three fields
    cpu_string                     = data.get("cpu_string", "")
    cpu_type, cpu_series, cpu_freq = parse_cpu_info(cpu_string)

    # Parse RAM
    ram_slots               = data.get("ram_slots", [])
    ram_config, _, ram_type = parse_ram_info(ram_slots)
    total_ram_clean         = clean_total_ram(ram_config)

    now      = datetime.now()
    date_str = f"{now.month}/{now.day:02d}/{now.year}"

    battery_health  = str(data.get("battery_health", "")).strip()
    condition       = str(data.get("condition", "")).strip()
    condition_grade = str(data.get("condition_grade", "A-Grade (Like-New)")).strip()

    if battery_health.lower() in {"", "unknown", "unavailable", "none", "n/a"}:
        battery_prefix = "Battery dead."
    else:
        battery_prefix = f"Battery Health: {battery_health}."

    battery_text = f"{battery_prefix} {condition}".strip()

    csad_value = data.get("csad_value", "").strip()

    # Write to columns
    new_row[0]  = csad_value if csad_value else "*"   # A
    new_row[1]  = date_str                            # B
    new_row[7]  = data.get("serial", "")             # H
    new_row[10] = battery_text                        # K
    new_row[11] = condition_grade                     # L
    new_row[18] = cpu_series                          # S  e.g. "8250U", "155H", "N4020"
    new_row[27] = ram_config                          # AB
    new_row[28] = total_ram_clean                     # AC
    new_row[29] = ram_type                            # AD
    new_row[40] = cpu_freq                            # AO
    new_row[41] = cpu_type                            # AP e.g. "Intel Core i5", "AMD Ryzen 7"

    daily_ws.append(new_row)

    # Ensure long condition notes in column K wrap instead of overflowing.
    row_idx = daily_ws.max_row
    daily_ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True)

    daily_wb.save(DAILY_FILE)

# ---------------------------------------------------
# ROUTES
# ---------------------------------------------------

@app.route("/ping", methods=["GET"])
def ping():
    """Test connectivity -- laptop uses this to verify server is reachable."""
    return jsonify({"status": "ok"}), 200


@app.route("/log", methods=["POST"])
def log_laptop():
    """Receive laptop data and append to Excel log."""
    data = request.get_json()

    if not data:
        return jsonify({"status": "error", "message": "No data received"}), 400

    required = ["model", "serial", "cpu_string", "ram_slots", "battery_health"]
    missing  = [f for f in required if f not in data]
    if missing:
        return jsonify({"status": "error", "message": f"Missing fields: {missing}"}), 400

    try:
        append_to_log(data)
        model = data.get("model", "unknown")
        print(f"  [{datetime.now().strftime('%H:%M:%S')}] Logged: {model} (CSAD: {data.get('csad_value', 'none')})")
        return jsonify({"status": "ok", "message": "Logged successfully"}), 200

    except ValueError as e:
        # Model not found in template
        return jsonify({"status": "error", "message": str(e)}), 404

    except Exception as e:
        print(f"  ERROR: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# ---------------------------------------------------
# MAIN
# ---------------------------------------------------

if __name__ == "__main__":
    print("\n" + "="*50)
    print("  LaptopSync Server")
    print("="*50)

    setup_directory()
    initialize_daily_log()

    print(f"\n  Server running on port {PORT}")
    print(f"  Laptops should point to this PC's IP address")
    print(f"  Press Ctrl+C to stop\n")
    print("="*50 + "\n")

    app.run(host="0.0.0.0", port=PORT, debug=False)